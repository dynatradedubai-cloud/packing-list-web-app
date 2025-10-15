
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

st.set_page_config(page_title="Packing List Generator", layout="wide")
st.title("📦 Packing List Generator")

uploaded_file = st.file_uploader("Upload Dump Excel File", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file, engine="openpyxl")

    # Rename columns to match expected names
    column_mapping = {
        'QUANTITY': 'QTY',
        'CRTNWEIGHT': 'CRTN WEIGHT'
    }
    df.rename(columns=column_mapping, inplace=True)

    required_columns = ['CARTONNO', 'PARTNO', 'QTY', 'REF1', 'PARTDESC', 'WEIGHT', 'MANFPART', 'CRTN WEIGHT', 'Brand']
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        st.error(f"Missing columns in uploaded file: {', '.join(missing_cols)}")
    else:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Packing List"

        # Styles
        center = Alignment(horizontal="center", vertical="center")
        left_middle = Alignment(horizontal="left", vertical="center")
        right_middle = Alignment(horizontal="right", vertical="center")
        bottom_left = Alignment(horizontal="left", vertical="bottom")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)

        # Header
        ws.merge_cells("A1:J1")
        ws["A1"] = "PACKING LIST"
        ws["A1"].alignment = center
        ws["A1"].font = bold_font

        ws.merge_cells("A2:C2")
        ws.merge_cells("A3:C3")
        ws.merge_cells("E2:J2")
        ws["E2"] = f"Invoice Number: {', '.join(df['REF1'].dropna().unique().astype(str))}"
        ws["E2"].alignment = left_middle

        ws.merge_cells("E3:J3")
        ws["E3"] = f"Date: {datetime.now().strftime('%d/%m/%Y')}"
        ws["E3"].alignment = left_middle

        # Table headers
        headers = ["Sl. No", "CARTONNO", "PARTNO", "QTY", "REF1", "PARTDESC", "WEIGHT", "MANFPART", "CRTN WEIGHT", "Brand"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = bold_font
            cell.alignment = center
            cell.border = thin_border

        # Group by CARTONNO and CRTN WEIGHT
        grouped = df.groupby(['CARTONNO', 'CRTN WEIGHT'], sort=False)
        start_row = 6
        sl_no = 1
        for (carton, weight), group in grouped:
            rows = group.shape[0]
            for i, (_, row) in enumerate(group.iterrows()):
                ws.cell(row=start_row + i, column=3, value=row['PARTNO'])
                ws.cell(row=start_row + i, column=4, value=row['QTY'])
                ws.cell(row=start_row + i, column=5, value=row['REF1'])
                ws.cell(row=start_row + i, column=6, value=row['PARTDESC'])
                ws.cell(row=start_row + i, column=7, value=row['WEIGHT'])
                ws.cell(row=start_row + i, column=8, value=row['MANFPART'])
                ws.cell(row=start_row + i, column=10, value=row['Brand'])

            # Merge Sl. No, CARTONNO, CRTN WEIGHT
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + rows - 1, end_column=1)
            ws.cell(row=start_row, column=1, value=sl_no)
            ws.cell(row=start_row, column=1).alignment = center

            ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + rows - 1, end_column=2)
            ws.cell(row=start_row, column=2, value=carton)
            ws.cell(row=start_row, column=2).alignment = center

            ws.merge_cells(start_row=start_row, start_column=9, end_row=start_row + rows - 1, end_column=9)
            ws.cell(row=start_row, column=9, value=weight)
            ws.cell(row=start_row, column=9).alignment = center

            start_row += rows
            sl_no += 1

        # Footer
        total_qty = df['QTY'].sum()
        total_weight = df['CRTN WEIGHT'].drop_duplicates().sum()
        package_count = df['CARTONNO'].nunique()

        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
        ws.cell(row=start_row, column=2, value="Total Quantity")
        ws.cell(row=start_row, column=2).alignment = right_middle
        ws.cell(row=start_row, column=4, value=total_qty)

        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=8)
        ws.cell(row=start_row, column=6, value="TOTAL CARTON WEIGHT")
        ws.cell(row=start_row, column=6).alignment = right_middle
        ws.cell(row=start_row, column=9, value=total_weight)

        start_row += 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
        ws.cell(row=start_row, column=1, value=f"NO OF PACKAGES : {package_count}")

        start_row += 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
        ws.cell(row=start_row, column=1, value=f"TOTAL GROSS WEIGHT : {round(total_weight)} KG")

        start_row += 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 3, end_column=10)
        ws.cell(row=start_row, column=1, value="AUTHORISED SIGNATORY")
        ws.cell(row=start_row, column=1).alignment = bottom_left

        # Apply borders to all used cells
        for row in ws.iter_rows(min_row=1, max_row=start_row + 3, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        st.success("Packing List created successfully!")

        st.download_button(
            label="Download Packing List",
            data=output.getvalue(),
            file_name="Packing_List.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
